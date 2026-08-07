from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_FILE = "upstream_test_data_template.xlsx"

HEADERS = [
    "Category",
    "API Name",
    "Test Case Name",
    "Request Direction",
    "Test Objective",
    "Method",
    "Endpoint",
    "Environment",
    "Headers",
    "Request Parameters",
    "Request Example",
    "Expected Status Code",
    "Expected Response",
    "Response Example",
    "Business Validation Points",
    "Signature Required",
    "Authentication Required",
    "Related Account",
    "Asset/Currency",
    "Amount",
    "Whitelisted Address/Entity",
    "Remarks",
]

ROWS = [
    ["Account Creation", "Account Creation - Successful", "Create account successfully", "Request/Response", "Verify that an account can be created successfully", "POST", "", "UAT/Test", "", "", "", "200/201", "Account is created successfully and returns account identifier, status, and related details", "", "Account ID is unique; status is correct; all mandatory fields are returned", "Yes/No", "Yes/No", "", "", "", "", "If applicable"],
    ["Account Creation", "Account Creation - Duplicate Request", "Create duplicate account", "Request/Response", "Verify handling of duplicate account creation", "POST", "", "UAT/Test", "", "", "", "400/409", "System returns duplicate or conflict error", "", "Error code and error message are clear and accurate", "Yes/No", "Yes/No", "", "", "", "", "If applicable"],
    ["Account Creation", "Account Creation - Missing Required Fields", "Missing required parameters", "Request/Response", "Verify validation when required parameters are missing", "POST", "", "UAT/Test", "", "", "", "400", "System returns parameter validation error", "", "Missing field information is clearly indicated", "Yes/No", "Yes/No", "", "", "", "", "If applicable"],
    ["Account Creation", "Account Creation - Invalid Parameters", "Invalid parameter format", "Request/Response", "Verify validation for invalid parameter formats", "POST", "", "UAT/Test", "", "", "", "400", "System returns invalid parameter error", "", "Field format validation works as expected", "Yes/No", "Yes/No", "", "", "", "", "If applicable"],
    ["Whitelisting", "Whitelisting - Successful Addition", "Add whitelist successfully", "Request/Response", "Verify that a whitelist entry can be added successfully", "POST", "", "UAT/Test", "", "", "", "200/201", "Whitelist entry is created successfully and returns whitelist ID/status", "", "Whitelist status, address/entity information, and returned data are correct", "Yes/No", "Yes/No", "", "", "", "To be filled", ""],
    ["Whitelisting", "Whitelisting - Duplicate Entry", "Add duplicate whitelist entry", "Request/Response", "Verify handling of duplicate whitelist entries", "POST", "", "UAT/Test", "", "", "", "400/409", "System returns duplicate or conflict error", "", "Error code and error message are clear and accurate", "Yes/No", "Yes/No", "", "", "", "To be filled", ""],
    ["Whitelisting", "Whitelisting - Invalid Address or Entity", "Invalid whitelist data", "Request/Response", "Verify validation for invalid whitelist address or entity", "POST", "", "UAT/Test", "", "", "", "400", "System returns validation error", "", "Address format or entity format validation works correctly", "Yes/No", "Yes/No", "", "", "", "To be filled", ""],
    ["Whitelisting", "Whitelisting - Status Query", "Query whitelist status", "Request/Response", "Verify whitelist status query", "GET", "", "UAT/Test", "", "", "", "200", "System returns whitelist status successfully", "", "Returned status matches actual configuration", "Yes/No", "Yes/No", "", "", "", "To be filled", ""],
    ["Fund Transfer", "Deposit - Successful", "Successful deposit", "Request/Response", "Verify successful deposit handling", "POST/WEBHOOK", "", "UAT/Test", "", "", "", "200", "Deposit is processed successfully and returns transaction ID/status", "", "Amount, asset/currency, account, and transaction status are correct", "Yes/No", "Yes/No", "To be filled", "To be filled", "To be filled", "", ""],
    ["Fund Transfer", "Deposit - Duplicate Notification", "Duplicate deposit notification", "Request/Response", "Verify idempotent handling of duplicate deposit notifications", "POST/WEBHOOK", "", "UAT/Test", "", "", "", "200", "Duplicate notification does not create duplicate credit entry", "", "Idempotency handling works correctly", "Yes/No", "Yes/No", "To be filled", "To be filled", "To be filled", "", ""],
    ["Fund Transfer", "Deposit - Invalid Amount", "Abnormal deposit amount", "Request/Response", "Verify handling of invalid or abnormal deposit amount", "POST/WEBHOOK", "", "UAT/Test", "", "", "", "400/422", "System returns amount validation error or business rejection", "", "Rejection reason is clear and accurate", "Yes/No", "Yes/No", "To be filled", "To be filled", "To be filled", "", ""],
    ["Fund Transfer", "Withdrawal - Successful", "Successful withdrawal", "Request/Response", "Verify successful withdrawal handling", "POST", "", "UAT/Test", "", "", "", "200", "Withdrawal request is accepted successfully and returns order ID/status", "", "Amount, fee, destination address, and status are correct", "Yes/No", "Yes/No", "To be filled", "To be filled", "To be filled", "To be filled", ""],
    ["Fund Transfer", "Withdrawal - Insufficient Balance", "Withdrawal with insufficient balance", "Request/Response", "Verify handling when balance is insufficient", "POST", "", "UAT/Test", "", "", "", "400/422", "System returns insufficient balance error", "", "Error code and message are clear and accurate", "Yes/No", "Yes/No", "To be filled", "To be filled", "To be filled", "To be filled", ""],
    ["Fund Transfer", "Withdrawal - Non-Whitelisted Address", "Withdrawal to non-whitelisted address", "Request/Response", "Verify rejection of withdrawal to a non-whitelisted address", "POST", "", "UAT/Test", "", "", "", "400/403", "System rejects the request due to whitelist restriction", "", "Whitelist validation is enforced correctly", "Yes/No", "Yes/No", "To be filled", "To be filled", "To be filled", "To be filled", ""],
    ["Fund Transfer", "Withdrawal - Risk Control Rejection", "Withdrawal rejected by risk control", "Request/Response", "Verify risk control rejection scenario", "POST", "", "UAT/Test", "", "", "", "403/422", "System returns risk control rejection", "", "Risk reason or rejection code is clear and accurate", "Yes/No", "Yes/No", "To be filled", "To be filled", "To be filled", "To be filled", ""],
    ["Webhook Response", "Webhook - Successful Callback", "Acknowledge successful callback", "Response Only", "Verify successful webhook reception and response", "POST", "", "UAT/Test", "", "", "", "200", "Receiver returns success acknowledgment as required", "", "Response format complies with upstream requirements; processing succeeds", "Yes/No", "Yes/No", "", "", "", "", ""],
    ["Webhook Response", "Webhook - Signature Verification Failure", "Invalid webhook signature", "Response Only", "Verify response when signature verification fails", "POST", "", "UAT/Test", "", "", "", "401/403", "Receiver returns signature verification failure", "", "Error response format complies with requirements", "Yes/No", "Yes/No", "", "", "", "", ""],
    ["Webhook Response", "Webhook - Missing Parameters", "Missing webhook parameters", "Response Only", "Verify response when required webhook parameters are missing", "POST", "", "UAT/Test", "", "", "", "400", "Receiver returns parameter validation error", "", "Missing field details are clearly indicated", "Yes/No", "Yes/No", "", "", "", "", ""],
    ["Webhook Response", "Webhook - Duplicate Callback", "Duplicate webhook callback", "Response Only", "Verify idempotent response for duplicate callback", "POST", "", "UAT/Test", "", "", "", "200", "Duplicate callback still returns success or idempotent result", "", "Business logic is not processed repeatedly", "Yes/No", "Yes/No", "", "", "", "", ""],
    ["Webhook Response", "Webhook - Internal System Error", "System exception during webhook processing", "Response Only", "Verify response when internal system error occurs", "POST", "", "UAT/Test", "", "", "", "500", "Receiver returns internal server error", "", "Error scenario aligns with retry mechanism expectations", "Yes/No", "Yes/No", "", "", "", "", ""],
]


def autosize_columns(ws):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            text = "" if value is None else str(value)
            widths[idx] = max(widths.get(idx, 0), min(len(text) + 2, 50))

    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    ws.append(HEADERS)
    for row in ROWS:
        ws.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

    ws.freeze_panes = "A2"
    autosize_columns(ws)

    wb.save(OUTPUT_FILE)
    print(OUTPUT_FILE)


if __name__ == "__main__":
    main()
