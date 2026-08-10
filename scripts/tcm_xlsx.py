#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""tcm_xlsx.py —— 测试用例管理模块（TCM）的 xlsx 导入 / 导出桥接脚本。

调用约定与 ``tmp/export_report_docx.py`` 保持一致：

    python scripts/tcm_xlsx.py <mode> <payloadPath>

其中 ``mode`` 为 ``export`` 或 ``import``，``payloadPath`` 指向一个临时 JSON 文件。
脚本读取 payload、执行任务、把结果写回 ``payload["outputPath"]``（导出为 .xlsx 二进制、
导入为 .json 结果文件），并在 stdout 打印产物绝对路径。

依赖缺失（未安装 openpyxl）时：
    - 向 stderr 打印约定错误标记 ``TCM_XLSX_ERROR:OPENPYXL_MISSING``
    - 以退出码 ``3`` 结束
    server.js 侦测到该标记 / 退出码后会自动降级为 CSV 并提示用户。

其他运行期错误统一打印 ``TCM_XLSX_ERROR:<CODE> <message>`` 并以非零退出码结束。
"""

from __future__ import annotations

import base64
import json
import os
import sys
import traceback
from typing import Any, Dict, List

# 约定错误码（与 server.js 中的常量保持一致，勿随意改动）
# 其中 BAD_ARGS / BAD_PAYLOAD / EMPTY_DATA / BAD_FILE 属于「调用方输入有误」，
# server.js 会把它们映射为 HTTP 4xx；RUNTIME 才是真正的服务端故障（5xx）。
ERR_OPENPYXL_MISSING = "OPENPYXL_MISSING"
ERR_BAD_ARGS = "BAD_ARGS"
ERR_BAD_PAYLOAD = "BAD_PAYLOAD"
ERR_EMPTY_DATA = "EMPTY_DATA"
ERR_BAD_FILE = "BAD_FILE"
ERR_RUNTIME = "RUNTIME"

EXIT_OK = 0
EXIT_RUNTIME = 1
EXIT_BAD_ARGS = 2
EXIT_OPENPYXL_MISSING = 3

# 列宽上下限（字符数），避免超长步骤把列撑爆
MIN_COLUMN_WIDTH = 8
MAX_COLUMN_WIDTH = 60


def fail(code: str, message: str, exit_code: int = EXIT_RUNTIME) -> None:
    """打印约定错误标记并退出。

    Args:
        code: 约定错误码。
        message: 人类可读的错误说明。
        exit_code: 进程退出码。
    """
    sys.stderr.write("TCM_XLSX_ERROR:%s %s\n" % (code, message))
    sys.stderr.flush()
    sys.exit(exit_code)


def require_openpyxl():
    """惰性导入 openpyxl；缺失时按约定降级信号退出。

    Returns:
        已导入的 ``openpyxl`` 模块。
    """
    try:
        import openpyxl  # noqa: F401  (运行期依赖，按需导入)
    except ImportError as error:  # pragma: no cover - 取决于运行环境
        fail(
            ERR_OPENPYXL_MISSING,
            "未安装 openpyxl（%s），请改用 CSV 导入导出或执行 pip install openpyxl" % error,
            EXIT_OPENPYXL_MISSING,
        )
    return openpyxl


def read_payload(payload_path: str) -> Dict[str, Any]:
    """读取并解析临时 payload JSON。

    Args:
        payload_path: payload 文件绝对路径。

    Returns:
        解析后的字典。
    """
    if not payload_path or not os.path.isfile(payload_path):
        fail(ERR_BAD_ARGS, "payload 文件不存在：%s" % payload_path, EXIT_BAD_ARGS)

    try:
        with open(payload_path, "r", encoding="utf-8") as handle:
            data = json.load(handle)
    except (OSError, ValueError) as error:
        fail(ERR_BAD_PAYLOAD, "payload 解析失败：%s" % error)
        return {}

    if not isinstance(data, dict):
        fail(ERR_BAD_PAYLOAD, "payload 顶层必须是对象")
    return data


def normalize_cell(value: Any) -> str:
    """把任意 JSON 值转成适合写入单元格的字符串。

    Args:
        value: 原始值。

    Returns:
        单元格文本；None 转为空串，布尔转「是 / 否」。
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (list, tuple)):
        return ",".join(normalize_cell(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def compute_column_width(header: str, cells: List[str]) -> float:
    """按内容估算列宽（中文按 2 个字符宽度计）。

    Args:
        header: 表头文本。
        cells: 该列所有单元格文本。

    Returns:
        列宽（openpyxl 字符宽度单位）。
    """

    def visual_width(text: str) -> int:
        width = 0
        for char in text.splitlines()[0] if text else "":
            width += 2 if ord(char) > 0x2E80 else 1
        return width

    widest = visual_width(header)
    for cell in cells[:200]:  # 只采样前 200 行，避免大表卡顿
        widest = max(widest, visual_width(cell))
    return float(max(MIN_COLUMN_WIDTH, min(MAX_COLUMN_WIDTH, widest + 2)))


def do_export(payload: Dict[str, Any]) -> str:
    """导出 xlsx。

    payload 约定：
        columns: [{key, label}]  —— 列定义（label 作为表头）
        rows: [[cell, ...]]      —— 已由前端 / 服务端拍平的行矩阵
        sheetName: str           —— 可选，工作表名
        title: str               —— 可选，冻结表头上方的标题（当前仅用于 sheet 名兜底）
        outputPath: str          —— 产物落盘绝对路径

    Args:
        payload: payload 字典。

    Returns:
        产物绝对路径。
    """
    openpyxl = require_openpyxl()
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    output_path = str(payload.get("outputPath") or "").strip()
    if not output_path:
        fail(ERR_BAD_PAYLOAD, "缺少 outputPath")

    columns = payload.get("columns")
    if not isinstance(columns, list) or not columns:
        fail(ERR_BAD_PAYLOAD, "缺少 columns")

    headers: List[str] = []
    for column in columns:
        if isinstance(column, dict):
            headers.append(normalize_cell(column.get("label") or column.get("key")))
        else:
            headers.append(normalize_cell(column))

    raw_rows = payload.get("rows")
    if not isinstance(raw_rows, list):
        fail(ERR_BAD_PAYLOAD, "缺少 rows")
    if not raw_rows:
        fail(ERR_EMPTY_DATA, "没有可导出的用例")

    rows: List[List[str]] = []
    for raw_row in raw_rows:
        if isinstance(raw_row, dict):
            # 兼容对象行：按 columns 的 key 取值
            row = []
            for column in columns:
                key = column.get("key") if isinstance(column, dict) else column
                row.append(normalize_cell(raw_row.get(key)))
            rows.append(row)
            continue
        cells = list(raw_row) if isinstance(raw_row, (list, tuple)) else [raw_row]
        row = [normalize_cell(cell) for cell in cells]
        # 补齐 / 截断到列数
        if len(row) < len(headers):
            row.extend([""] * (len(headers) - len(row)))
        rows.append(row[: len(headers)])

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = (str(payload.get("sheetName") or "").strip() or "测试用例")[:31]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4F6BED")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.append(headers)
    for index in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for row in rows:
        sheet.append(row)

    for row_index in range(2, len(rows) + 2):
        for column_index in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.alignment = body_align
            cell.border = border

    for column_index, header in enumerate(headers, start=1):
        column_cells = [row[column_index - 1] for row in rows]
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = compute_column_width(header, column_cells)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), len(rows) + 1)

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    workbook.save(output_path)
    return output_path


def do_import(payload: Dict[str, Any]) -> str:
    """解析上传的 xlsx，把 ``{ok, headers, rows}`` 写入 outputPath（JSON）。

    payload 约定：
        contentBase64: str  —— xlsx 原始二进制的 base64
        fileName: str       —— 原始文件名（仅用于错误提示）
        sheetName: str      —— 可选，指定工作表；缺省取第一个
        maxRows: int        —— 可选，最多解析行数（默认 5000）
        outputPath: str     —— 结果 JSON 落盘路径

    Args:
        payload: payload 字典。

    Returns:
        结果 JSON 的绝对路径。
    """
    openpyxl = require_openpyxl()
    import io

    output_path = str(payload.get("outputPath") or "").strip()
    if not output_path:
        fail(ERR_BAD_PAYLOAD, "缺少 outputPath")

    content_base64 = str(payload.get("contentBase64") or "")
    if not content_base64:
        fail(ERR_BAD_PAYLOAD, "缺少 contentBase64")

    file_name = str(payload.get("fileName") or "upload.xlsx")
    max_rows = payload.get("maxRows")
    max_rows = int(max_rows) if isinstance(max_rows, (int, float)) and max_rows > 0 else 5000

    try:
        binary = base64.b64decode(content_base64, validate=False)
    except Exception as error:  # noqa: BLE001 - base64 解码失败原因多样
        fail(ERR_BAD_PAYLOAD, "文件内容解码失败（%s）：%s" % (file_name, error))
        return ""

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(binary), data_only=True, read_only=True)
    except Exception as error:  # noqa: BLE001 - openpyxl 抛出的异常类型不固定
        # 用户上传的文件损坏 / 根本不是 xlsx，属于输入问题而非服务端故障，
        # 必须用 BAD_FILE 让 server.js 回 4xx，否则前端会误报「服务异常，请重试」。
        fail(ERR_BAD_FILE, "无法读取 xlsx（%s）：%s" % (file_name, error))
        return ""

    sheet_name = str(payload.get("sheetName") or "").strip()
    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    headers: List[str] = []
    rows: List[Dict[str, str]] = []
    truncated = False

    for row_index, raw_row in enumerate(sheet.iter_rows(values_only=True)):
        if row_index == 0:
            raw_headers = [normalize_cell(cell).strip() for cell in raw_row]
            # 去掉尾部完全空白的列（Excel 常见的幽灵列）
            while raw_headers and not raw_headers[-1]:
                raw_headers.pop()
            for column_index, text in enumerate(raw_headers):
                headers.append(text if text else "列%d" % (column_index + 1))
            continue

        if len(rows) >= max_rows:
            truncated = True
            break

        record: Dict[str, str] = {}
        has_value = False
        for column_index, header in enumerate(headers):
            value = normalize_cell(raw_row[column_index]).strip() if column_index < len(raw_row) else ""
            record[header] = value
            if value:
                has_value = True
        if has_value:
            rows.append(record)

    workbook.close()

    if not headers:
        fail(ERR_EMPTY_DATA, "xlsx 首行没有表头：%s" % file_name)

    result = {
        "ok": True,
        "fileName": file_name,
        "sheetName": sheet.title,
        "headers": headers,
        "rows": rows,
        "rowCount": len(rows),
        "truncated": truncated,
    }

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as handle:
        json.dump(result, handle, ensure_ascii=False)
    return output_path


def main() -> None:
    """入口：解析参数并分发到 export / import。"""
    if len(sys.argv) < 3:
        fail(ERR_BAD_ARGS, "用法：python scripts/tcm_xlsx.py <export|import> <payloadPath>", EXIT_BAD_ARGS)

    mode = str(sys.argv[1] or "").strip().lower()
    payload_path = sys.argv[2]

    if mode not in ("export", "import"):
        fail(ERR_BAD_ARGS, "未知模式：%s（仅支持 export / import）" % mode, EXIT_BAD_ARGS)

    payload = read_payload(payload_path)

    try:
        output_path = do_export(payload) if mode == "export" else do_import(payload)
    except SystemExit:
        raise
    except Exception as error:  # noqa: BLE001 - 兜底，保证 server 端能拿到原因
        sys.stderr.write(traceback.format_exc())
        fail(ERR_RUNTIME, "执行失败：%s" % error)
        return

    sys.stdout.write(output_path + "\n")
    sys.stdout.flush()
    sys.exit(EXIT_OK)


if __name__ == "__main__":
    main()
